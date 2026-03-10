from datetime import datetime
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import pandas as pd


def export_workbook(sheets: dict):
    filename = f"export_axonaut_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

            ws = writer.book[sheet_name[:31]]

            # ligne d'entête en gras
            for cell in ws[1]:
                cell.font = Font(bold=True)

            # filtre auto
            ws.auto_filter.ref = ws.dimensions

            # freeze première ligne
            ws.freeze_panes = "A2"

            # largeur auto simple
            for col_idx, column_cells in enumerate(ws.columns, start=1):
                max_length = 0
                for cell in column_cells:
                    try:
                        value = "" if cell.value is None else str(cell.value)
                        if len(value) > max_length:
                            max_length = len(value)
                    except Exception:
                        pass
                adjusted_width = min(max_length + 2, 40)
                ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    print(f"Fichier généré : {filename}")